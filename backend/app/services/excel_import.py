"""Parse the AG Grow Excel order form into catalog rows.

The Excel has repeating sections. Each section starts with a header row whose
first cell is exactly "S.NO". That row's remaining non-empty cells (after
"S.NO", "INSECTICIDES"/"HERBICIDE"/"FUNGICIDE"/"BIO STIMULANTS", and
"PACKING TYPE") are the size labels for the section. The *second* cell on that
header row names the category (e.g. INSECTICIDES). Following rows until the
next header or a blank row are product rows. A cell with "X" (any case) under a
size column marks that packing as **not** manufactured in-house (third-party /
listed only); all other cells in that size column—including empty—are treated as
sizes you manufacture and thus become orderable in the app after import.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

_CATEGORY_CANONICAL = {
    "INSECTICIDES": "Insecticides",
    "INSECTICIDE": "Insecticides",
    "HERBICIDES": "Herbicide",
    "HERBICIDE": "Herbicide",
    "FUNGICIDES": "Fungicide",
    "FUNGICIDE": "Fungicide",
    "BIO STIMULANTS": "Bio Stimulants",
    "BIOSTIMULANTS": "Bio Stimulants",
    "BIO STIMULANT": "Bio Stimulants",
}


@dataclass
class ParsedProduct:
    s_no: int
    name: str
    packing_type: str | None
    available_sizes: list[str]


@dataclass
class ParsedGroup:
    category: str
    label: str
    column_headers: list[str]
    products: list[ParsedProduct] = field(default_factory=list)


def _norm(v) -> str:
    return ("" if v is None else str(v)).strip()


def _is_header_row(row: tuple) -> bool:
    return _norm(row[0]).upper() == "S.NO"


def parse_excel(path: str | Path) -> list[ParsedGroup]:
    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active
    groups: list[ParsedGroup] = []
    current: ParsedGroup | None = None

    for row in ws.iter_rows(values_only=True):
        if not any(_norm(c) for c in row):
            current = None
            continue

        if _is_header_row(row):
            cat_raw = _norm(row[1]).upper()
            category = _CATEGORY_CANONICAL.get(cat_raw, _norm(row[1]).title() or "Uncategorized")
            headers: list[str] = []
            for cell in row[3:]:
                label = _norm(cell)
                if label:
                    headers.append(label)
            label = f"{category} — {' / '.join(headers)}" if headers else category
            current = ParsedGroup(category=category, label=label, column_headers=headers)
            groups.append(current)
            continue

        if current is None:
            continue

        s_no_raw = _norm(row[0])
        if not s_no_raw.isdigit():
            continue
        name = _norm(row[1])
        if not name:
            continue
        packing_type = _norm(row[2]) or None

        available: list[str] = []
        for idx, size in enumerate(current.column_headers):
            cell_idx = 3 + idx
            if cell_idx >= len(row):
                break
            raw = _norm(row[cell_idx]).upper()
            if raw != "X":
                available.append(size)

        current.products.append(
            ParsedProduct(
                s_no=int(s_no_raw),
                name=name,
                packing_type=packing_type,
                available_sizes=available,
            )
        )

    return groups


def group_by_category(groups: Iterable[ParsedGroup]) -> dict[str, list[ParsedGroup]]:
    out: dict[str, list[ParsedGroup]] = {}
    for g in groups:
        out.setdefault(g.category, []).append(g)
    return out
